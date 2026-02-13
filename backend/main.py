import io
import re
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook
from database import get_db, init_db

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup():
    init_db()


# --- helpers ---

def categorize(description: str, rules: list[tuple[str, str]]) -> str:
    upper = description.upper()
    for keyword, category in rules:
        if keyword.upper() in upper:
            return category
    return "Other"


def parse_excel(file_bytes: bytes, filename: str, rules: list[tuple[str, str]]) -> list[dict]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    transactions = []
    current_card = None
    in_transactions = False

    for row in rows:
        col0 = str(row[0] or "").strip()
        col1 = str(row[1] or "").strip()

        if re.match(r"^\d{6}\*{6}\d{4}$", col0) and col1:
            current_card = col1
            in_transactions = False
            continue

        if col0 == "Datum" and col1 == "Bokfört":
            in_transactions = True
            continue

        if col0 == "Totalt belopp":
            in_transactions = False
            continue

        if not row[0] and not row[2]:
            continue

        if in_transactions and row[0] and row[2]:
            date_val = row[0]
            if hasattr(date_val, "strftime"):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val)

            booked_val = row[1]
            if hasattr(booked_val, "strftime"):
                booked_str = booked_val.strftime("%Y-%m-%d")
            else:
                booked_str = str(booked_val or "")

            description = str(row[2]).strip()
            city = str(row[3] or "").strip()
            currency = str(row[4] or "").strip()
            amount = float(row[6] or 0)

            transactions.append({
                "date": date_str,
                "booked_date": booked_str,
                "description": description,
                "city": city,
                "currency": currency,
                "amount": amount,
                "card_holder": current_card or "Unknown",
                "category": categorize(description, rules),
                "source_file": filename,
            })

    return transactions


# --- endpoints ---

@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx files supported")

    data = await file.read()
    db = get_db()

    rules = db.execute("SELECT keyword, category FROM category_rules").fetchall()
    rules = [(r["keyword"], r["category"]) for r in rules]

    txns = parse_excel(data, file.filename, rules)

    inserted = 0
    for t in txns:
        try:
            db.execute(
                """INSERT INTO transactions (date, booked_date, description, city, currency, amount, card_holder, category, source_file)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (t["date"], t["booked_date"], t["description"], t["city"],
                 t["currency"], t["amount"], t["card_holder"], t["category"], t["source_file"]),
            )
            inserted += 1
        except Exception:
            pass  # dedup constraint

    db.commit()
    db.close()
    return {"inserted": inserted, "total_in_file": len(txns)}


@app.get("/transactions")
def get_transactions(category: str = None, card_holder: str = None, source_file: str = None):
    db = get_db()
    query = "SELECT * FROM transactions WHERE 1=1"
    params = []
    if category:
        query += " AND category = ?"
        params.append(category)
    if card_holder:
        query += " AND card_holder = ?"
        params.append(card_holder)
    if source_file:
        query += " AND source_file = ?"
        params.append(source_file)
    query += " ORDER BY date DESC"
    rows = db.execute(query, params).fetchall()
    db.close()
    return [dict(r) for r in rows]


@app.get("/categories")
def get_categories():
    db = get_db()
    rows = db.execute("SELECT name FROM categories ORDER BY name").fetchall()
    db.close()
    return [r["name"] for r in rows]


class CategoryCreate(BaseModel):
    name: str


@app.post("/categories")
def create_category(body: CategoryCreate):
    db = get_db()
    try:
        db.execute("INSERT INTO categories (name) VALUES (?)", (body.name,))
        db.commit()
    except Exception:
        db.close()
        raise HTTPException(409, "Category already exists")
    db.close()
    return {"ok": True}


@app.delete("/categories/{name:path}")
def delete_category(name: str):
    if name in ("Other", "Excluded"):
        raise HTTPException(400, "Cannot delete built-in category")
    db = get_db()
    db.execute("UPDATE transactions SET category = 'Other' WHERE category = ?", (name,))
    db.execute("DELETE FROM category_rules WHERE category = ?", (name,))
    db.execute("DELETE FROM categories WHERE name = ?", (name,))
    db.commit()
    db.close()
    return {"ok": True}


@app.get("/category-rules")
def get_rules():
    db = get_db()
    rows = db.execute("SELECT keyword, category FROM category_rules ORDER BY keyword").fetchall()
    db.close()
    return [dict(r) for r in rows]


class CategoryUpdate(BaseModel):
    category: str


@app.put("/transactions/{txn_id}/category")
def update_transaction_category(txn_id: int, body: CategoryUpdate):
    db = get_db()
    db.execute("UPDATE transactions SET category = ?, manual_category = 1 WHERE id = ?", (body.category, txn_id))
    db.execute("INSERT OR IGNORE INTO categories (name) VALUES (?)", (body.category,))
    db.commit()
    db.close()
    return {"updated": 1}


class RuleUpdate(BaseModel):
    category: str


@app.put("/category-rules/{keyword:path}")
def update_rule(keyword: str, body: RuleUpdate):
    db = get_db()
    # Upsert rule
    existing = db.execute("SELECT id FROM category_rules WHERE keyword = ?", (keyword,)).fetchone()
    if existing:
        db.execute("UPDATE category_rules SET category = ? WHERE keyword = ?", (body.category, keyword))
    else:
        db.execute("INSERT INTO category_rules (keyword, category) VALUES (?, ?)", (keyword, body.category))

    # Ensure category exists
    db.execute("INSERT OR IGNORE INTO categories (name) VALUES (?)", (body.category,))

    # Update matching transactions, but skip ones manually categorized
    escaped = keyword.upper().replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
    updated = db.execute(
        "UPDATE transactions SET category = ? WHERE UPPER(description) LIKE ? ESCAPE '\\' AND manual_category = 0",
        (body.category, f"%{escaped}%"),
    ).rowcount

    db.commit()
    db.close()
    return {"updated": updated}


class RuleCreate(BaseModel):
    keyword: str
    category: str


@app.post("/category-rules")
def create_rule(body: RuleCreate):
    db = get_db()
    try:
        db.execute("INSERT INTO category_rules (keyword, category) VALUES (?, ?)", (body.keyword, body.category))
    except Exception:
        db.close()
        raise HTTPException(409, "Rule already exists for this keyword")

    db.execute("INSERT OR IGNORE INTO categories (name) VALUES (?)", (body.category,))

    escaped = body.keyword.upper().replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
    updated = db.execute(
        "UPDATE transactions SET category = ? WHERE UPPER(description) LIKE ? ESCAPE '\\' AND manual_category = 0",
        (body.category, f"%{escaped}%"),
    ).rowcount

    db.commit()
    db.close()
    return {"updated": updated}


@app.get("/source-files")
def get_source_files():
    db = get_db()
    rows = db.execute("SELECT DISTINCT source_file FROM transactions ORDER BY source_file").fetchall()
    db.close()
    return [r["source_file"] for r in rows]
